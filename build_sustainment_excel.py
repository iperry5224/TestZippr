"""
Build SAELAR & SOPRA Sustainment Plan - Wow Factor Edition
Creates a visually stunning Excel workbook with tabs, shading, and grading.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Color palette - openpyxl needs aRGB (FF + 6-char hex)
def _rgb(h):
    return "FF" + h if len(h) == 6 else h

COLORS = {
    "header_dark": _rgb("1E3A5F"),
    "header_accent": _rgb("2E5A8C"),
    "saelar": _rgb("0D47A1"),
    "sopra": _rgb("1565C0"),
    "ec2": _rgb("37474F"),
    "security": _rgb("B71C1C"),
    "availability": _rgb("2E7D32"),
    "updates": _rgb("F9A825"),
    "documentation": _rgb("6A1B9A"),
    "monitoring": _rgb("00838F"),
    "compliance": _rgb("4A148C"),
    "daily": _rgb("E8F5E9"),
    "weekly": _rgb("E3F2FD"),
    "monthly": _rgb("FFF8E1"),
    "quarterly": _rgb("FCE4EC"),
    "semi_annual": _rgb("F3E5F5"),
    "annual": _rgb("E0F2F1"),
    "alt_row": _rgb("F5F5F5"),
    "white": _rgb("FFFFFF"),
    "text_dark": _rgb("212121"),
    "text_light": _rgb("FFFFFF"),
}

def style_header(ws, row, cols, fill_hex, font_color="FFFFFF", bold=True):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        cell.font = Font(color=font_color, bold=bold, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_cell(ws, row, col, fill_hex=None, font_color="212121", wrap=True):
    cell = ws.cell(row=row, column=col)
    if fill_hex:
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    cell.font = Font(color=font_color, size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)

def category_fill(cat):
    m = {"Security": COLORS["security"], "Availability": COLORS["availability"],
         "Updates": COLORS["updates"], "Documentation": COLORS["documentation"],
         "Monitoring": COLORS["monitoring"], "Compliance": COLORS["compliance"]}
    return m.get(cat, COLORS["alt_row"])

def frequency_fill(freq):
    m = {"Daily": COLORS["daily"], "Weekly": COLORS["weekly"], "Monthly": COLORS["monthly"],
         "Quarterly": COLORS["quarterly"], "Semi-annual": COLORS["semi_annual"],
         "Annual": COLORS["annual"], "Per audit": COLORS["quarterly"],
         "After deploy": COLORS["monthly"], "Per activity": COLORS["weekly"]}
    return m.get(freq, COLORS["white"])

def platform_fill(sid):
    if sid.startswith("SAE"): return COLORS["saelar"]
    if sid.startswith("SOP"): return COLORS["sopra"]
    return COLORS["ec2"]

wb = Workbook()
wb.remove(wb.active)

# ========== TAB 1: Overview (Dashboard-style) ==========
ws_overview = wb.create_sheet("📊 Overview", 0)
ws_overview.sheet_properties.tabColor = COLORS["header_dark"]
ws_overview.column_dimensions["A"].width = 18
ws_overview.column_dimensions["B"].width = 45
ws_overview.row_dimensions[1].height = 35

# Title
ws_overview.merge_cells("A1:B1")
ws_overview["A1"] = "SAELAR & SOPRA Sustainment Plan"
ws_overview["A1"].fill = PatternFill(start_color=COLORS["header_dark"], end_color=COLORS["header_dark"], fill_type="solid")
ws_overview["A1"].font = Font(color="FFFFFF", bold=True, size=16)
ws_overview["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Deployment info blocks
overview_data = [
    ("SAELAR", "Port: 8484 (EC2) / 8443 standalone | https://saelar.ngrok.dev"),
    ("SOPRA", "Port: 8080 | https://sopra.ngrok.dev"),
    ("EC2", "ubuntu@18.232.122.255 | Key: saelar-sopra-key.pem | Path: /opt/apps/"),
]
for i, (label, val) in enumerate(overview_data):
    r = 3 + i * 2
    ws_overview[f"A{r}"] = label
    ws_overview[f"B{r}"] = val
    fill = platform_fill("SAE" if "SAELAR" in label else "SOP" if "SOPRA" in label else "EC2")
    ws_overview[f"A{r}"].fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    ws_overview[f"A{r}"].font = Font(color="FFFFFF", bold=True)
    ws_overview[f"B{r}"].fill = PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid")
    ws_overview[f"B{r}"].alignment = Alignment(wrap_text=True)

# Schedule at a glance
ws_overview["A9"] = "Schedule at a Glance"
ws_overview["A9"].font = Font(bold=True, size=12)
ws_overview.merge_cells("A10:B10")
ws_overview["A10"] = "Daily → Process health, ngrok | Weekly → Backup, Log review, Disk | Monthly → Dep audit, OS patches"
ws_overview["A10"].fill = PatternFill(start_color=COLORS["daily"], end_color=COLORS["daily"], fill_type="solid")
ws_overview.merge_cells("A11:B11")
ws_overview["A11"] = "Quarterly → Credential rotation, Python deps, SSH key | Semi-annual → Streamlit, Bedrock | Annual → Restore test, AI audit, Python"
ws_overview["A11"].fill = PatternFill(start_color=COLORS["quarterly"], end_color=COLORS["quarterly"], fill_type="solid")

# ========== TAB 2: Schedule (Frequency Grading) ==========
ws_sched = wb.create_sheet("📅 Schedule", 1)
ws_sched.sheet_properties.tabColor = COLORS["availability"]
headers = ["Frequency", "Activities", "IDs", "Owner"]
for c, h in enumerate(headers, 1):
    cell = ws_sched.cell(row=1, column=c, value=h)
    cell.fill = PatternFill(start_color=COLORS["header_accent"], end_color=COLORS["header_accent"], fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

schedule_rows = [
    ("Daily", "Process health (SAELAR, SOPRA), ngrok tunnel", "SAE-SUS-03, SOP-SUS-03, EC2-SUS-03", "Ops"),
    ("Weekly", "Backup, Log review, Disk space", "SAE-SUS-04, SOP-SUS-04, SAE-SUS-09, SOP-SUS-10, EC2-SUS-04", "Ops"),
    ("Monthly", "Dependency audit, OS patches", "SAE-SUS-02, SOP-SUS-02, EC2-SUS-01", "Dev/Ops"),
    ("Quarterly", "Credential rotation, Python deps, SSH key", "SAE-SUS-01, SOP-SUS-01, SAE-SUS-05, SOP-SUS-05, EC2-SUS-02", "Infra/Dev"),
    ("Semi-annual", "Streamlit, Bedrock models", "SAE-SUS-06, SOP-SUS-06, SOP-SUS-07", "Dev"),
    ("Annual", "Backup restore test, AI data minimization, Python version", "SOP-SUS-12, EC2-SUS-05", "Security/Dev"),
]
for r, row in enumerate(schedule_rows, 2):
    fill = frequency_fill(row[0])
    for c, val in enumerate(row, 1):
        cell = ws_sched.cell(row=r, column=c, value=val)
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
ws_sched.column_dimensions["A"].width = 14
ws_sched.column_dimensions["B"].width = 50
ws_sched.column_dimensions["C"].width = 45
ws_sched.column_dimensions["D"].width = 12

# ========== TAB 3: Deploy Reference ==========
ws_deploy = wb.create_sheet("🚀 Deploy", 2)
ws_deploy.sheet_properties.tabColor = COLORS["updates"]
ws_deploy["A1"] = "SOPRA Deploy to EC2"
ws_deploy["A1"].font = Font(bold=True, size=14)
ws_deploy["A1"].fill = PatternFill(start_color=COLORS["sopra"], end_color=COLORS["sopra"], fill_type="solid")
ws_deploy["A1"].font = Font(color="FFFFFF", bold=True, size=14)
ws_deploy.merge_cells("A1:D1")
deploy_steps = [
    "1. scp -i saelar-sopra-key.pem sopra_setup.py sopra_controls.py ubuntu@18.232.122.255:/tmp/sopra_update/",
    "2. scp -i saelar-sopra-key.pem -r sopra demo_csv_data ubuntu@18.232.122.255:/tmp/sopra_update/",
    "3. ssh -i saelar-sopra-key.pem ubuntu@18.232.122.255",
    "4. Run: pkill -f 'streamlit run sopra_setup'; sleep 2; cp -f /tmp/sopra_update/*.py /opt/apps/; rm -rf /opt/apps/sopra; cp -r /tmp/sopra_update/sopra /tmp/sopra_update/demo_csv_data /opt/apps/; cd /opt/apps; nohup venv/bin/streamlit run sopra_setup.py --server.port 8080 --server.address 0.0.0.0 --server.headless true > /tmp/sopra.log 2>&1 &",
]
for i, step in enumerate(deploy_steps, 3):
    ws_deploy[f"A{i}"] = step
    ws_deploy[f"A{i}"].fill = PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid")
    ws_deploy[f"A{i}"].alignment = Alignment(wrap_text=True)
ws_deploy.column_dimensions["A"].width = 100

# ========== TAB 4: SAELAR Activities ==========
ws_saelar = wb.create_sheet("🛡️ SAELAR", 3)
ws_saelar.sheet_properties.tabColor = COLORS["saelar"]
saelar_headers = ["ID", "Category", "Metric", "Activity", "Acceptance Criteria", "Steps", "Expected Result", "Frequency", "Owner", "Last Done", "Next Due", "Pass/Fail", "Notes"]
for c, h in enumerate(saelar_headers, 1):
    cell = ws_saelar.cell(row=1, column=c, value=h)
    cell.fill = PatternFill(start_color=COLORS["saelar"], end_color=COLORS["saelar"], fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

saelar_data = [
    ("SAE-SUS-01", "Security", "Credential Rotation", "Rotate AWS credentials", "New creds configured; old revoked", "Generate IAM keys; update config; restart; verify; revoke old", "SAELAR operational", "Quarterly", "Infra", "", "", "", ""),
    ("SAE-SUS-02", "Security", "Dependency Audit", "Scan Python deps for vulns", "No critical/high CVEs", "pip audit; review; update; re-test", "Deps current", "Monthly", "Dev", "", "", "", ""),
    ("SAE-SUS-03", "Availability", "Process Health", "Verify SAELAR process on EC2", "Process active; port 8484", "SSH; ps aux | grep nist_setup; curl localhost:8484", "Process running", "Daily", "Ops", "", "", "", ""),
    ("SAE-SUS-04", "Availability", "Backup", "Backup assessment data", "Backup stored; restore tested", "Copy /opt/apps/; backup JSON; store offsite", "Backup complete", "Weekly", "Ops", "", "", "", ""),
    ("SAE-SUS-05", "Updates", "Python Dependencies", "Update pip packages", "requirements.txt updated; tests pass", "pip list --outdated; update; UAT; deploy", "Deps current", "Quarterly", "Dev", "", "", "", ""),
    ("SAE-SUS-06", "Updates", "Streamlit Version", "Keep Streamlit current", "Streamlit compatible", "Check release notes; update venv; smoke test; deploy", "Streamlit stable", "Semi-annual", "Dev", "", "", "", ""),
    ("SAE-SUS-07", "Documentation", "Runbook Update", "Runbook reflects deploy steps", "Runbook matches EC2 layout", "Review SOPRA_EC2_DEPLOY.md; execute; update drift", "Runbook accurate", "After deploy", "Dev", "", "", "", ""),
    ("SAE-SUS-08", "Documentation", "Change Log", "Log sustainment changes", "Changes documented", "Update CHANGELOG; note version date action", "Change log current", "Per activity", "Owner", "", "", "", ""),
    ("SAE-SUS-09", "Monitoring", "Log Review", "Review app logs for errors", "No critical errors", "tail /tmp/saelar.log; search ERROR/CRITICAL", "Logs reviewed", "Weekly", "Ops", "", "", "", ""),
    ("SAE-SUS-10", "Compliance", "Audit Trail", "Retain assessment history", "Traceable", "Verify log retention; export/backup audit data", "Audit trail present", "Per audit", "Compliance", "", "", "", ""),
]
for r, row in enumerate(saelar_data, 2):
    for c, val in enumerate(row, 1):
        cell = ws_saelar.cell(row=r, column=c, value=val)
        if c == 2:  # Category
            cell.fill = PatternFill(start_color=category_fill(val), end_color=category_fill(val), fill_type="solid")
            cell.font = Font(color="FFFFFF", size=9)
        elif c == 8:  # Frequency
            cell.fill = PatternFill(start_color=frequency_fill(val), end_color=frequency_fill(val), fill_type="solid")
        else:
            fill = COLORS["white"] if r % 2 == 0 else COLORS["alt_row"]
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
ws_saelar.column_dimensions["A"].width = 12
ws_saelar.column_dimensions["B"].width = 14
ws_saelar.column_dimensions["C"].width = 18
ws_saelar.column_dimensions["D"].width = 22
ws_saelar.column_dimensions["E"].width = 28
ws_saelar.column_dimensions["F"].width = 35
ws_saelar.column_dimensions["G"].width = 18
ws_saelar.column_dimensions["H"].width = 12
ws_saelar.column_dimensions["I"].width = 10
ws_saelar.column_dimensions["J"].width = 11
ws_saelar.column_dimensions["K"].width = 11
ws_saelar.column_dimensions["L"].width = 10
ws_saelar.column_dimensions["M"].width = 20

# ========== TAB 5: SOPRA Activities ==========
ws_sopra = wb.create_sheet("📋 SOPRA", 4)
ws_sopra.sheet_properties.tabColor = COLORS["sopra"]
for c, h in enumerate(saelar_headers, 1):
    cell = ws_sopra.cell(row=1, column=c, value=h)
    cell.fill = PatternFill(start_color=COLORS["sopra"], end_color=COLORS["sopra"], fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

sopra_data = [
    ("SOP-SUS-01", "Security", "Credential Rotation", "Rotate AWS credentials", "New creds; Bedrock AI works", "Generate IAM keys; update config; restart; verify AI; revoke old", "SOPRA operational", "Quarterly", "Infra", "", "", "", ""),
    ("SOP-SUS-02", "Security", "Dependency Audit", "Scan Python deps for vulns", "No critical/high CVEs", "pip audit; review; update; re-test", "Deps current", "Monthly", "Dev", "", "", "", ""),
    ("SOP-SUS-03", "Availability", "Process Health", "Verify SOPRA process on EC2", "Process active; port 8080", "SSH; ps aux | grep sopra_setup; curl localhost:8080", "Process running", "Daily", "Ops", "", "", "", ""),
    ("SOP-SUS-04", "Availability", "Backup", "Backup assessment data and JSON", "Backup stored; restore tested", "Copy /opt/apps/; backup poam.json incidents.json; store offsite", "Backup complete", "Weekly", "Ops", "", "", "", ""),
    ("SOP-SUS-05", "Updates", "Python Dependencies", "Update pip packages", "requirements.txt updated; tests pass", "pip list --outdated; update; UAT; deploy", "Deps current", "Quarterly", "Dev", "", "", "", ""),
    ("SOP-SUS-06", "Updates", "Streamlit Version", "Keep Streamlit current", "Streamlit compatible", "Check release notes; update venv; smoke test; deploy", "Streamlit stable", "Semi-annual", "Dev", "", "", "", ""),
    ("SOP-SUS-07", "Updates", "Bedrock Model Availability", "Verify Bedrock models enabled", "AI features work", "Check Bedrock console; update ai_engine.py if deprecated; re-test", "AI functional", "Semi-annual", "Dev", "", "", "", ""),
    ("SOP-SUS-08", "Documentation", "Runbook Update", "Runbook reflects deploy steps", "Runbook matches EC2 layout", "Review SOPRA_EC2_DEPLOY.md; execute; update drift", "Runbook accurate", "After deploy", "Dev", "", "", "", ""),
    ("SOP-SUS-09", "Documentation", "Change Log", "Log sustainment changes", "Changes documented", "Update CHANGELOG; note version date action", "Change log current", "Per activity", "Owner", "", "", "", ""),
    ("SOP-SUS-10", "Monitoring", "Log Review", "Review app logs for errors", "No critical errors", "tail /tmp/sopra.log; search ERROR/CRITICAL", "Logs reviewed", "Weekly", "Ops", "", "", "", ""),
    ("SOP-SUS-11", "Compliance", "Audit Readiness", "Output artifacts audit-ready", "POA&M Risk SSP formats valid", "Generate samples; review ATO checklist; update templates", "Output audit-ready", "Per audit", "Compliance", "", "", "", ""),
    ("SOP-SUS-12", "Compliance", "AI Data Minimization", "Confirm AI sends only aggregate data", "No PII/evidence to Bedrock", "Review ai_engine.py; verify payloads; document flow", "Data minimization verified", "Annual", "Security", "", "", "", ""),
]
for r, row in enumerate(sopra_data, 2):
    for c, val in enumerate(row, 1):
        cell = ws_sopra.cell(row=r, column=c, value=val)
        if c == 2:
            cell.fill = PatternFill(start_color=category_fill(val), end_color=category_fill(val), fill_type="solid")
            cell.font = Font(color="FFFFFF", size=9)
        elif c == 8:
            cell.fill = PatternFill(start_color=frequency_fill(val), end_color=frequency_fill(val), fill_type="solid")
        else:
            fill = COLORS["white"] if r % 2 == 0 else COLORS["alt_row"]
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
for col in "ABCDEFGHIJKLM":
    ws_sopra.column_dimensions[col].width = ws_saelar.column_dimensions[col].width

# ========== TAB 6: EC2 Shared ==========
ws_ec2 = wb.create_sheet("☁️ EC2 Shared", 5)
ws_ec2.sheet_properties.tabColor = COLORS["ec2"]
for c, h in enumerate(saelar_headers, 1):
    cell = ws_ec2.cell(row=1, column=c, value=h)
    cell.fill = PatternFill(start_color=COLORS["ec2"], end_color=COLORS["ec2"], fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ec2_data = [
    ("EC2-SUS-01", "Security", "OS Patches", "Apply Ubuntu security updates", "Critical/high patches applied", "sudo apt update && apt upgrade -y; reboot if kernel; verify apps", "OS patched", "Monthly", "Ops", "", "", "", ""),
    ("EC2-SUS-02", "Security", "SSH Key", "Rotate or verify EC2 key pair", "Key valid; access confirmed", "Verify key; test SSH; update if rotated", "SSH works", "Quarterly", "Infra", "", "", "", ""),
    ("EC2-SUS-03", "Availability", "ngrok Tunnel", "Verify ngrok tunnels active", "sopra.ngrok.dev and saelar.ngrok.dev resolve", "curl both URLs; restart ngrok if needed", "Tunnels active", "Daily", "Ops", "", "", "", ""),
    ("EC2-SUS-04", "Availability", "Disk Space", "Monitor disk usage", "/opt/apps and /tmp have space", "df -h; review log growth; rotate logs", "Disk adequate", "Weekly", "Ops", "", "", "", ""),
    ("EC2-SUS-05", "Updates", "Python Version", "Ensure Python supported", "Python 3.10+; venv compatible", "python3 --version; check EOL; plan upgrade", "Python supported", "Annual", "Dev", "", "", "", ""),
]
for r, row in enumerate(ec2_data, 2):
    for c, val in enumerate(row, 1):
        cell = ws_ec2.cell(row=r, column=c, value=val)
        if c == 2:
            cell.fill = PatternFill(start_color=category_fill(val), end_color=category_fill(val), fill_type="solid")
            cell.font = Font(color="FFFFFF", size=9)
        elif c == 8:
            cell.fill = PatternFill(start_color=frequency_fill(val), end_color=frequency_fill(val), fill_type="solid")
        else:
            fill = COLORS["white"] if r % 2 == 0 else COLORS["alt_row"]
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
for col in "ABCDEFGHIJKLM":
    ws_ec2.column_dimensions[col].width = ws_saelar.column_dimensions[col].width

# Freeze panes
for ws in [ws_saelar, ws_sopra, ws_ec2]:
    ws.freeze_panes = "A2"

out_path = os.path.join(os.path.dirname(__file__), "SAELAR_SOPRA_Sustainment_Plan.xlsx")
wb.save(out_path)
print(f"Created: {out_path}")

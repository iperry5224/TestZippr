"""
Create professionally styled UAT spreadsheets for SAELAR and SOPRA.
Output: .xlsx files compatible with Google Sheets (File > Import).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import csv
from pathlib import Path

# ---- Color palette ----
COLORS = {
    "title_bg": "0F172A",      # Slate 900
    "header_bg": "1E40AF",    # Blue 800
    "header_col1": "1E40AF",
    "header_col2": "1D4ED8",
    "header_col3": "2563EB",
    "category_colors": {
        "Security": "7F1D1D",       # Red 900
        "Relevance": "14532D",      # Green 900
        "Ease of Use": "4C1D95",    # Violet 900
        "Functionality": "1E3A8A",  # Blue 900
        "Performance": "713F12",     # Amber 900
        "Compliance": "134E4A",     # Teal 900
    },
    "alt_row1": "FFFFFF",
    "alt_row2": "F1F5F9",
    "row_accent": "E0E7FF",
    "border": "94A3B8",
}

def load_uat_csv(filepath):
    """Load UAT data from CSV."""
    rows = []
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
    return rows

def style_sheet(ws, title, subtitle, num_data_cols, data_end_row):
    """Apply global styling: freeze panes, column widths, auto-filter."""
    ws.freeze_panes = "A4"  # Freeze header rows
    widths = [12, 12, 14, 22, 45, 50, 55, 40, 10, 25, 12, 12]
    for i, w in enumerate(widths[:num_data_cols], 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 50)
    if data_end_row >= 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(num_data_cols)}{data_end_row}"

def add_title_block(ws, title, subtitle, platform_color):
    """Add fancy title block at top."""
    ws.merge_cells("A1:K1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color=platform_color, end_color=platform_color, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:K2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=11, color="FFFFFF", italic=True)
    ws["A2"].fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

def add_header_row(ws, headers, row_num):
    """Add styled header row with alternating column shading."""
    shades = [COLORS["header_col1"], COLORS["header_col2"], COLORS["header_col3"]]
    pass_fail_col = 9  # Pass/Fail column (after No Source Req'd)
    no_source_col = 2  # No Source Req'd - highlight as black-box tests
    thin_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="medium", color="FFFFFF"),
        bottom=Side(style="medium", color="FFFFFF"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        shade = "15803D" if col == pass_fail_col else ("166534" if col == no_source_col else shades[(col - 1) % 3])
        cell.fill = PatternFill(start_color=shade, end_color=shade, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row_num].height = 32

def add_data_rows(ws, data_rows, start_row):
    """Add data rows with category-based accent and alternating shading."""
    cat_colors = COLORS["category_colors"]
    prev_cat = None
    use_alt = False
    for i, row in enumerate(data_rows):
        r = start_row + i
        cat = row[2] if len(row) > 2 else ""  # Category is 3rd column (after No Source Req'd)
        if cat != prev_cat:
            prev_cat = cat
            use_alt = not use_alt
        fill_color = COLORS["alt_row2"] if use_alt else COLORS["alt_row1"]
        accent = cat_colors.get(cat, "64748B")
        row_border = Border(
            left=Side(style="medium", color=accent),
            right=Side(style="thin", color=COLORS["border"]),
            top=Side(style="thin", color=COLORS["border"]),
            bottom=Side(style="thin", color=COLORS["border"]),
        )
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = row_border
            if col == 1:
                cell.font = Font(size=10, bold=True, color="1E293B")
            if col == 2:  # No Source Req'd - green accent for black-box tests
                cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                cell.font = Font(size=10, bold=True, color="166534")
        txt_len = len(str(row[4] or "")) + len(str(row[5] or "")) if len(row) > 5 else 0
        ws.row_dimensions[r].height = max(28, min(72, txt_len // 35 + 24))

def add_pass_fail_validation(ws, col_letter, start_row, end_row):
    """Add dropdown for Pass/Fail column."""
    dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,N/A,Blocked"',
        allow_blank=True
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)

def add_legend(ws, start_row):
    """Add legend / instructions with styling."""
    r = start_row
    ws.cell(row=r, column=1, value="📋 LEGEND & INSTRUCTIONS")
    ws.cell(row=r, column=1).font = Font(bold=True, size=14, color="1E40AF")
    r += 1
    ws.cell(row=r, column=1, value="• Pass/Fail: Select from dropdown (Pass, Fail, N/A, Blocked)")
    r += 1
    ws.cell(row=r, column=1, value="• Category colors (left border): Security=Red | Relevance=Green | Ease of Use=Violet | Functionality=Blue | Performance=Amber | Compliance=Teal")
    r += 1
    ws.cell(row=r, column=1, value="• Use header filter arrows to sort/filter by Category, Pass/Fail, or any column")
    r += 1
    ws.cell(row=r, column=1, value="• Freeze panes: Row 3 (headers) stay visible when scrolling")
    r += 1
    ws.cell(row=r, column=1, value="• No Source Req'd (✓): All tests are black-box — run via UI, network inspection (browser DevTools), or CLI. No application source code access needed.")
    r += 1

def create_combined_uat():
    """Create single workbook with SAELAR and SOPRA as separate tabs."""
    wb = Workbook()
    base = Path(__file__).parent

    # SAELAR tab (first sheet - rename active)
    ws_saelar = wb.active
    ws_saelar.title = "SAELAR"
    data = load_uat_csv(base / "UAT_SAELAR.csv")
    headers, rows = data[0], data[1:]
    end_row = 4 + len(rows) - 1
    add_title_block(ws_saelar, "🛡️ SAELAR — User Acceptance Test Plan",
                    "Security Architecture & Evaluation | NIST 800-53 Rev 5 | Cloud Assessment Platform",
                    COLORS["title_bg"])
    add_header_row(ws_saelar, headers, 3)
    add_data_rows(ws_saelar, rows, 4)
    add_pass_fail_validation(ws_saelar, "I", 4, end_row)
    add_legend(ws_saelar, end_row + 2)
    style_sheet(ws_saelar, "SAELAR", "", len(headers), end_row)

    # SOPRA tab (second sheet)
    ws_sopra = wb.create_sheet(title="SOPRA")
    data = load_uat_csv(base / "UAT_SOPRA.csv")
    headers, rows = data[0], data[1:]
    end_row = 4 + len(rows) - 1
    add_title_block(ws_sopra, "🛡️ SOPRA — User Acceptance Test Plan",
                    "SAE On-Premise Risk Assessment | 200 Controls | ISSO Toolkit | AI-Powered Automation",
                    "0F172A")
    add_header_row(ws_sopra, headers, 3)
    add_data_rows(ws_sopra, rows, 4)
    add_pass_fail_validation(ws_sopra, "I", 4, end_row)
    add_legend(ws_sopra, end_row + 2)
    style_sheet(ws_sopra, "SOPRA", "", len(headers), end_row)

    out = base / "UAT_SAELAR_SOPRA_Combined.xlsx"
    wb.save(out)
    print(f"Created: {out}")
    print("  - Tab 1: SAELAR (23 test cases)")
    print("  - Tab 2: SOPRA (25 test cases)")
    return out

if __name__ == "__main__":
    create_combined_uat()
    print("\nSingle file with two tabs. Import into Google Sheets via File > Import > Upload.")
